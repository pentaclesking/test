import os
import shutil
import time
import traceback
import openpyxl
from imap_tools import MailBox, AND, OR, NOT, A, H, U
import constants
import telebot
import datetime
from zipfile import ZipFile

xls_path = constants.xls_path
temp_path = constants.temp_path
bot = telebot.TeleBot(constants.token)


def repack_xls(file):
    # Создаем временную папку
    os.makedirs(temp_path, exist_ok=True)

    # Распаковываем excel как zip в нашу временную папку
    with ZipFile(xls_path + file) as excel_container:
        excel_container.extractall(temp_path)

    # Переименовываем файл с неверным названием
    wrong_file_path = os.path.join(temp_path, 'xl', 'SharedStrings.xml')
    correct_file_path = os.path.join(temp_path, 'xl', 'sharedStrings.xml')
    os.rename(wrong_file_path, correct_file_path)

    # Запаковываем excel обратно в zip и переименовываем в исходный файл
    shutil.make_archive('yourfile', 'zip', temp_path)
    # os.rename('yourfile.zip', 'yourfile.xlsx')
    shutil.move('yourfile.zip', xls_path + file)

    time.sleep(5)
    shutil.rmtree(temp_path, ignore_errors=True)
    time.sleep(0.5)
    # time.sleep(2)
    # shutil.rmtree(temp_path)


def send_message(message):
    bot.send_message(constants.admin_tg, message)


def is_already_processed(filename):
    with open(constants.log, mode='r', encoding='utf-8') as f:
        line_list = [line.rstrip('\n') for line in f]
    return True if filename in line_list else False


def add_to_log(filename):
    with open(constants.log, mode='a', encoding='utf-8') as f:
        f.write(filename + '\n')


def save_attachments_only(fop, from_dt):
    with MailBox(constants.fops[fop]['server']).login(constants.fops[fop]['login'], constants.fops[fop]['password'], 'INBOX') as mailbox:
        for msg in mailbox.fetch(AND(date_gte=datetime.date(from_dt.year, from_dt.month, from_dt.day))):
            # for msg in mailbox.fetch(AND(subject='Рожко', date_gte=dt.date(2022, 1, 15))):
            for att in msg.attachments:
                print(att.filename, att.content_type)
                filename = att.filename.replace(':', '').replace('/', '').replace('\\', '')
                with open(xls_path + filename, 'wb') as f:
                    f.write(att.payload)


def parse_xls(sh, first, fop):
    lines = []
    for i in range(first, sh.max_row + 1):
        if str(sh.cell(i, constants.npp_clmn).value).isdecimal():
            lines.append(f'{sh.cell(i, 9).value};{constants.fops[fop]["kassa"]};{sh.cell(i, 4).value};'
                         f'{sh.cell(i, 5).value};{sh.cell(i, 3).value};{constants.edrpou_np}\n')
    if lines:
        with open(constants.out_path + 'file.csv', 'a', encoding='utf-8') as f:
            f.writelines(lines)
    return True


def find_first_row(sh):
    for i in range(1, sh.max_row + 1):
        if sh.cell(i, constants.npp_clmn).value == 1:
            return i


def check_columns(sh, row):
    for column in constants.get_clmn:
        if sh.cell(row, column).value != constants.get_clmn[column]:
            return False
    return True


def remove_all_files(dir):
    for file in os.listdir(dir):
        os.remove(dir + file)
    # shutil.rmtree(dir)
    # os.makedirs(dir, exist_ok=True)


def set_work_dir():
    abspath = os.path.abspath(__file__)
    dname = os.path.dirname(abspath)
    os.chdir(dname)


def process_file(fop, file):
    repack_xls(file)

    wb = openpyxl.load_workbook(xls_path + file)
    sh = wb.active
    for items in sorted(sh.merged_cells.ranges):
        sh.unmerge_cells(str(items))

    first = find_first_row(sh)
    if not first:
        send_message(f'{__file__}\nНе найден первый ряд в файле {file}')
        raise Exception

    if not check_columns(sh, first - 2):
        send_message(f'{__file__}\nНе пройдена проверка на столбцы в файле {file}')
        raise Exception

    if parse_xls(sh=sh, first=first, fop=fop):
        return True


def process_attachments(fop):
    os.makedirs(xls_path, exist_ok=True)
    remove_all_files(xls_path)

    from_date = datetime.date.today() - datetime.timedelta(constants.days_ago)
    with MailBox(constants.fops[fop]['server']).login(constants.fops[fop]['login'], constants.fops[fop]['password'], 'INBOX') as mailbox:
        for msg in mailbox.fetch(AND(date_gte=datetime.date(from_date.year, from_date.month, from_date.day))):
            for att in msg.attachments:
                filename = att.filename.replace(':', '').replace('/', '').replace('\\', '')
                if fop.lower() not in filename.lower() or 'xls' not in filename.lower():
                    print(filename, '--- not reestr ---')
                    continue
                if is_already_processed(filename):
                    print(filename,  '--- already processed, skipping ---')
                else:
                    with open(xls_path + filename, 'wb') as f:
                        f.write(att.payload)
                    if process_file(fop, filename):
                        add_to_log(filename)
                        print(filename, '--- Successfully processed and added to log---')


def process_fop(fop):
    process_attachments(fop)


try:
    set_work_dir()
    for fop in constants.fops:
        print('Processing FOP', fop)
        process_fop(fop)
except Exception as e:
    send_message(f"Ошибка в программе {__file__}\n{str(e)}\n{traceback.format_exc()}")
    raise


# from_date = datetime.date.today() - datetime.timedelta(constants.days_ago)
# save_attachments_only('Zoho2', from_date)

